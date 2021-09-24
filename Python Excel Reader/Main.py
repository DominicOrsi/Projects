# Got working at 9/16/2021 12:25 AM
from openpyxl import Workbook, load_workbook
wb = load_workbook("9.13.21 DS Moves.xlsx")
ws = wb.active
rowCount = ws.max_row
colCount = ws.max_column

# make a function for this??
wn = []
sn = []
mainSN = []
mainWN = []

# adding data to wn and sn list respectivly
for i in range(1, rowCount + 1):
    for j in range(1, colCount + 1):
        data = ws.cell(row=i, column=j).value # gets the value of the cell and stores it as data
        if j == 1 and i > 1:
            sn.append(data)
        elif j == 2 and i > 1:
            wn.append(data)
        elif j == colCount - 1 and i > 1:
            mainSN.append(data)
        elif j == colCount and i > 1:
            mainWN.append(data)

# inits mulitiple to the size of mainSN
multiple = [0] * len(mainSN)

# print(mainSN) # commented out using (control + k + c), uncomment is (control + k + u)
# print(mainWN)
# print(multiple)

# removing none from the lists
# make a function to do this
noNoneWN = filter(None.__ne__, wn)
wn = list(noNoneWN)
noNoneSN = filter(None.__ne__, sn)
sn = list(noNoneSN)

# iterate through mainSN and compare its values against sn to see if there is any dups
for i in range(0, len(mainSN)):
    for j in range(0, len(sn)):
        # print("SN is:", sn[j], "\nMain SN is:", mainSN[i])
        if sn[j] == mainSN[i]:
            multiple[i] = multiple[i] + 1
            # print("We have a match\n")

ws.cell(row=1, column=colCount + 1, value="Total Moves")
# iterating through rowCount at 2 since we want the row entries to be at 2. To compensate we must add + 1 to rowCount.
# multiple[i - 2] is -2 because of i start at 2 for the row entries and in order to grab the data we must start at 0 and go to 17
# now adding a number value in the column next to WN to show how many times the WN appeared
for i in range(2, rowCount + 1):
    for j in range(colCount + 1, colCount + 2):
        ws.cell(row=i, column=j, value=multiple[i - 2])

# saves the WB to a new file
wb.save("DS Moves.xlsx")

print("Counted all moves up")