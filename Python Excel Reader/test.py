# import load_workbook
from openpyxl import load_workbook


wb = load_workbook("9.13.21 DS Moves.xlsx")
ws = wb.active

rowCount = ws.max_row
colCount = ws.max_column

print("Max row:", rowCount, "\nMax Column:", colCount)


wn = []
sp = []

'''
for i in range(1, rowCount + 1):
    for j in range(1, colCount + 1):
        data = ws.cell(row=i, column=j).value
        else:
            if j == 1 and i > 1:
                wn.append(data)
            elif j == 2 and i > 1:
                sp.append(data)
print(wn)
'''