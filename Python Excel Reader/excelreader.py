# import load_workbook
from openpyxl import load_workbook
# import webdriver, allows the uploading of files (in this case excel) to the web
from selenium import webdriver

# loading of the excel spreadsheet (e.g. June Moves)
wb =  load_workbook("9.13.21 DS Moves.xlsx")
# seting the sheet to active (e.g. 6/1/2021)
sheet = wb.active # wb["SheetName"]

rowCount =  sheet.max_row
columnCount = sheet.max_column
print("Max row:", rowCount, "\nMax Column:", columnCount)

sheet['C8'] = '= SUM(C2:C6) '

'''
start_col = 2 # 'C' column index
end_col = 3 # 'D' colunm index
# this only iterates through column C
for i in range(1, sheet.max_row + 1):
    row = [cell.value for cell in sheet[i] [start_col:end_col]]
    print(row)
'''
# what was I doing here???
# sheet.cell(row = rowCount + 2, colum = 3).value = 5

'''
# iterating through the excel sheet
for i in range(1, rowCount + 1):
    for j in range(1, columnCount + 1):
        data = sheet.cell(row=i, column=j).value
        print(data, end='   ')
    print('\n')
'''